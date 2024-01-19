from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import warnings, re, os
import argparse
warnings.filterwarnings(action='ignore')


class ExtractData():
    def __init__(self, dir_name: str):
        self.today = datetime.today().strftime('%Y-%m-%d')
        self.dir_name = dir_name
        self.file_path: str = f'input/{dir_name}'
        self.save_path: str = 'output'


    def get_postal_code(self, text: str):
        """
        우편번호 추출

        Args:
            text (str): 우편번호를 추출할 텍스트

        Returns:
            postal_code (str): 추출된 우편번호
        """        
        postal_code_pattern = re.compile(r'\((\d{5})\)') # 정규식 패턴
        match = postal_code_pattern.search(text)         # 패턴 매칭
        if match:
            postal_code = match.group(1)
        else: postal_code = '00000'

        return postal_code


    def run_all_file_monthly(self):
        """
        [monthly] 경로에 있는 모든 파일의 추출 실행
        """        
        file_list: list = os.listdir(self.file_path)
        file_list = [file for file in file_list if '.xlsx' in file]

        result_journal = pd.DataFrame()
        result_distribution_list = pd.DataFrame()

        for file in file_list:
            
            now_df = self.read_and_to_dataframe(file)
            
            date_ = file.split('.')[0].split('_')[-1]
            date_ = datetime.strptime(date_, '%Y%m%d')
            date_ = date_.strftime('%Y.%m.%d')
            
            temp_journal = self.extract_journal(now_df, date_)
            temp_distribution_list = self.extract_distribution_list(now_df)

            result_journal = pd.concat([result_journal, temp_journal])
            result_distribution_list = pd.concat([result_distribution_list, 
                                                temp_distribution_list])
            print("file name : ", file, 
                " / now file length : ", len(now_df), len(temp_journal), 
                " / total file length : ", len(result_journal))
           

        writer = pd.ExcelWriter(f'{self.save_path}/저널홍보제품목록_{self.dir_name}_{self.today}.xlsx', engine='openpyxl')
        
        result_journal.to_excel(writer, sheet_name='저널복붙')
        result_distribution_list.to_excel(writer, sheet_name='배포처목록')
        writer.save()
        print("Monthly Done!")


    def run_all_file_weekly(self):
        """
        [weekly] 경로에 있는 모든 파일의 추출 실행
        """        
        file_list: list = os.listdir(self.file_path)
        file_list = [file for file in file_list if '.xlsx' in file]

        result_copyright = pd.DataFrame()

        for file in file_list:
            
            now_df = self.read_and_to_dataframe(file)
    
            date_ = file.split('.')[0].split('_')[-1]
            date_ = datetime.strptime(date_, '%Y%m%d')
            date_ = date_.strftime('%Y.%m.%d')
        
                
            temp_copyright = self.extract_copyright(now_df, date_)
            result_copyright = pd.concat([result_copyright, temp_copyright])

            print("file name : ", file, 
                " / now file length : ", len(now_df), len(temp_copyright), 
                " / total file length : ", len(result_copyright))
           
            
        result_copyright.to_excel(f'{self.save_path}/저작권_{self.dir_name}_{self.today}.xlsx', engine='xlsxwriter')
        print("Weekly Done!")


    def count_target_rows(self, workbook, length: int):
        """
        추출 대상인 노란 색상으로 칠해진 셀의 개수 체크

        Args:
            workbook (workbook): openpyxl로 연 workbook
            length (int): 전체 DataFrame의 길이

        Returns:
            color_cnt (int): 노란 색상의 개수
        """        
        ## 색상 기준 추출 대상 row 찾기 ##
        ws = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
        pre_color_index = '00000000'
        color_cnt: int = 0
        for row in range(4, length+1):
            try:
                now_color = ws.cell(column=2, row=row).fill.fgColor.index
                if now_color == 43: 
                    color_cnt += 1
                    pre_color_index = ws.cell(column=2, row=row).fill.fgColor.index
                elif now_color == '00000000' and pre_color_index == 43:
                    color_cnt += 1
            except: pass
        return color_cnt


    def read_and_to_dataframe(self, file: str):
        """
        파일을 열고 데이터 프레임화

        Args:
            file (str): 대상 파일명

        Returns:
             df (pd.DataFrame): 정제한 데이터 프레임
        """        
        wb = load_workbook(f"{self.file_path}/{file}", data_only=True)
        df = pd.DataFrame(wb.active.values)
        
        color_cnt = self.count_target_rows(wb, len(df))

        df.columns = df.iloc[2,:]
        df = df.iloc[3:color_cnt+3, :]
        df['번호'].fillna(0, inplace=True)
        
    
        return df


    def extract_journal(self, now_df, date_):
        """
        [monthly] 추출 시트 1번인 '저널' 관련 결과물 추출        

        Args:
            now_df (pd.DataFrame): 작업할 대상 데이터프레임

        Returns:
            result_df (pd.DataFrame): 추출한 데이터프레임
        """        
        result_df = pd.DataFrame()
        col_list: list = ['인증번호\n(예정)', '인증일자', '제품명', 'GS번호', 'S/W분류', '제품설명', '회사명', '홈페이지', '시험원',
            '업무담당자', '전화번호', '이메일', '주소']  # 결과로 가져올 데이터 컬럼 리스트
        now_df['인증일자'] = date_
        result_df = now_df[col_list]
        

        return result_df


    def extract_distribution_list(self, now_df):
        """
        [monthly] 추출 시트 2번인 '배포처 목록' 관련 결과물 추출        

        Args:
            now_df (pd.DataFrame): 작업할 대상 데이터프레임

        Returns:
            result_df (pd.DataFrame): 추출한 데이터프레임
        """        
        result_df = pd.DataFrame()
        now_df['국문성명'] = now_df['업무담당자'].apply(lambda x: x.split()[0])
        now_df['직위'] = now_df['업무담당자'].apply(lambda x: x.split()[-1] if len(x.split()) == 2 else '-')
        now_df['우편번호'] = now_df['주소'].apply(lambda x: self.get_postal_code(x))

        # 필요한 데이터만 뽑기
        col_list: list = ['국문성명', '주소', '우편번호', '회사명', '직위', '이메일']
        result_df = now_df[col_list]

        return result_df


    def extract_copyright(self, now_df, date_):
        """
        [weekly] '저작권' 관련 결과물 추출        

        Args:
            now_df (pd.DataFrame): 작업할 대상 데이터프레임

        Returns:
            result_df (pd.DataFrame): 추출한 데이터프레임
        """
        
        result_df = pd.DataFrame
        col_list: list = ['인증일자', '인증번호\n(예정)', 'GS번호', '회사명', '제품명', '시험원']
        
        now_df['인증일자'] = date_
        result_df = now_df[col_list]
        result_df['시험원'].fillna('', inplace=True)
        result_df['시험원'] = result_df['시험원'].apply(lambda x: x.split(',')[0].split('\n')[0])
        
        return result_df          



if __name__=="__main__":
    dir_name = input("세부 경로를 입력해주세요. : ")
    parser = argparse.ArgumentParser()
    parser.add_argument("-t", "--type", type=str, default="weekly")
    args = parser.parse_args()

    extract = ExtractData(dir_name)    
    if args.type == 'weekly':
        extract.run_all_file_weekly()
    elif args.type == 'monthly':    
        extract.run_all_file_monthly()
    else:
        raise AttributeError