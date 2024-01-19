from openpyxl import load_workbook
import pandas as pd
import re
import datetime

class Transform():
    
    """
        출력 양식
        
        0. 분야 및 번호: 소프트웨어분야

        1. 인증로고 명칭: goodsoftware

        2. 인증번호: {00-0000}

        3. 모델명(제품설명) : {모델명} ({제품설명})

        4. 회사명(제조사)  /  홈페이지 주소 :  {회사명}/ {홈페이지 주소}

        5. 업체 담당자 연락처: {성함 및 직급}, {전화번호}, {이메일 주소}

        6. 제품 컬러 사진 (해상도 높은 것으로 주세요)
    """
    
    
    
    def __init__(self):
        
        # 경로 유동적으로 지정
        wb = load_workbook('/Users/geon-woo/Desktop/3.produce_hwp/input/저널홍보제품_추출목록.xlsx', data_only=True)
        df = pd.DataFrame(wb.active.values)
        ws = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        
        nowDate = datetime.datetime.now()
        f = open('/Users/geon-woo/Desktop/3.produce_hwp/output/새로운 인증 제품 목록_'+
                 nowDate.strftime("%Y-%m-%d") + ".txt", 'w')
        
        
        for i in range(2, len(df)+1):
            try:
                # 각 행의 필요한 셀 값을 저장
                gsNum = ws.cell(column=2, row=i).value
                modelName = ws.cell(column=4, row=i).value
                modelInfo = ws.cell(column=7, row=i).value
                compName = ws.cell(column=8, row=i).value
                compURL = ws.cell(column=9, row=i).value
                managerInfo = ws.cell(column=11, row=i).value
                managerPhone = ws.cell(column=12, row=i).value
                managerEmail = ws.cell(column=13, row=i).value
                # 회사 영문명, 특수문자, 오른쪽 공백 제거
                compName = (re.sub(r"[^\uAC00-\uD7A30㈜]", " ", compName)).rstrip()
                
                promotionForm = f"""
                새로운 인증 제품\n\n\n\n
0. 분야 및 번호: 소프트웨어분야\n
1. 인증로고 명칭: goodsoftware\n
2. 인증번호: {gsNum}\n
3. 모델명(제품설명) : {modelName} ({modelInfo})\n
4. 회사명(제조사)  /  홈페이지 주소 : {compName} / {compURL}\n
5. 업체 담당자 연락처: {managerInfo}, {managerPhone}, {managerEmail}\n
6. 제품 컬러 사진 (해상도 높은 것으로 주세요)\n\n\n\n\n\n\n\n\n\n
                """
                print(promotionForm, file=f)
            except:
                print(f"{i}번째 시도에서 에러 발생")
                pass            
        f.close()
        
    
if __name__ == "__main__":
    start = Transform()